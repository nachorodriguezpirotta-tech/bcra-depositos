#!/usr/bin/env python3
"""
Depósitos en USD del Sector Privado (BCRA) -> Excel con gráfico + cuadro.

Fuente: https://www.bcra.gob.ar/depositos-y-otros-pasivos-de-las-entidades-financieras/
Archivo: diar_dep.xls  (hoja "Sector_privado", columna CB = Depósitos USD del Sector Privado, en millones de u$s)

Genera 'depositos_usd_sector_privado.xlsx' con:
  - Hoja "Serie": serie diaria desde la asunción de Milei (10/12/2023).
  - Hoja "Hoja 1": gráfico de evolución + cuadro (monto hoy y fechas hacia atrás, Δ en USD y Δ en %).

Uso:
  python3 generate.py             # baja el archivo y genera el Excel
  python3 generate.py --no-download   # usa el .xls ya bajado en /tmp
"""
import sys, re, os, json, urllib.request
import datetime as dt

import xlrd
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.dates as mdates

URL = "https://www.bcra.gob.ar/archivos/Pdfs/PublicacionesEstadisticas/diar_dep.xls"
SRC = "/tmp/diar_dep.xls"
SHEET = "Sector_privado"
CB_COL = 79  # columna CB (0-indexed): Depósitos USD Sector Privado, total, en millones u$s
MILEI = dt.datetime(2023, 12, 10)  # asunción
OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "depositos_usd_sector_privado.xlsx")

# Puntos de comparación (replican el cuadro en crudo de Ignacio).
# kind: "today" | "days" (n días atrás) | "yearend" (31/12 del año anterior) | "milei"
LOOKBACKS = [
    ("Último", "today", None),
    ("1 Mes", "days", 30),
    ("Fin 2025", "yearend", None),
    ("1 Año", "days", 365),
    ("Inicio Milei", "milei", None),
]


def download():
    print(f"Descargando {URL} ...")
    req = urllib.request.Request(URL, headers={"User-Agent": "Mozilla/5.0"})
    with urllib.request.urlopen(req, timeout=180) as r, open(SRC, "wb") as f:
        f.write(r.read())
    print(f"  -> {SRC} ({os.path.getsize(SRC)/1e6:.0f} MB)")


# Money Market USD — FUENTE OFICIAL: CAFCI (planilla diaria).
# Gonzalo: sumar la columna P (índice 15) de la sección "Mercado de Dinero Dolar
# Estadounidense" (en la planilla son las filas ~P3960:P4096). Acá se hace robusto:
# se ubica el header de la sección y se suma hasta el próximo header.
CAFCI_URL = "https://api.pub.cafci.org.ar/pb_get"
CAFCI_SRC = "/tmp/cafci.xlsx"
MM_SECTION = "Mercado de Dinero Dolar Estadounidense"
MM_COL = 15  # columna P
# Fallback (si CAFCI no responde): argentinadatos
MM_FALLBACK_URL = "https://api.argentinadatos.com/v1/finanzas/fci/mercadoDinero/ultimo"
MM_FALLBACK_RE = re.compile(r"d[oó]lar|dolar|usd|u\$s", re.I)


def _fmt_fecha(f):
    if isinstance(f, str):
        p = f.split("/")
        if len(p) == 3 and len(p[2]) == 2:
            return f"{p[0]}/{p[1]}/20{p[2]}"
        return f
    try:
        return f.strftime("%d/%m/%Y")
    except Exception:
        return str(f) if f else ""


def fetch_mm_usd():
    """Patrimonio total de FCI Money Market en USD (millones de u$s).
    Fuente oficial CAFCI; fallback argentinadatos. Devuelve (fecha, total_millones) o None."""
    # 1) CAFCI oficial
    try:
        req = urllib.request.Request(CAFCI_URL, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=90) as r, open(CAFCI_SRC, "wb") as f:
            f.write(r.read())
        wb = openpyxl.load_workbook(CAFCI_SRC, data_only=True, read_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        start = next((i + 1 for i, row in enumerate(rows)
                      if isinstance(row[0], str) and row[0].strip() == MM_SECTION), None)
        if start is None:
            raise ValueError("sección MM USD no encontrada en planilla CAFCI")
        total = 0.0
        fecha = None
        for row in rows[start:]:
            a = row[0]
            if not (isinstance(a, str) and a.strip()):
                continue
            if "clase" not in a.lower():
                break  # arrancó la próxima sección
            v = row[MM_COL]
            if isinstance(v, (int, float)):
                total += v
            if fecha is None:
                fecha = row[4]  # col E = Fecha
        return _fmt_fecha(fecha), total / 1e6
    except Exception as e:
        print(f"  [aviso] CAFCI falló ({e}); uso argentinadatos como respaldo")
    # 2) fallback argentinadatos
    try:
        req = urllib.request.Request(MM_FALLBACK_URL, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=30) as r:
            data = json.load(r)
        usd = [f for f in data if MM_FALLBACK_RE.search(f.get("fondo", ""))]
        total = sum((f.get("patrimonio") or 0) for f in usd)
        fecha = usd[0].get("fecha") if usd else ""
        return _fmt_fecha(fecha), total / 1e6
    except Exception as e:
        print(f"  [aviso] no se pudo obtener Money Market USD: {e}")
        return None


def load_series():
    """Devuelve lista [(date, valor_millones_usd)] ordenada por fecha."""
    wb = xlrd.open_workbook(SRC, on_demand=True)
    sh = wb.sheet_by_name(SHEET)
    pat = re.compile(r"^\d{2}/\d{2}/\d{4}$")
    data = []
    for r in range(sh.nrows):
        a = sh.cell_value(r, 0)
        if isinstance(a, str) and pat.match(a):
            v = sh.cell_value(r, CB_COL)
            if isinstance(v, float):
                data.append((dt.datetime.strptime(a, "%d/%m/%Y"), v))
    data.sort()
    return data


def nearest_on_or_before(data, target):
    c = [d for d in data if d[0] <= target]
    return c[-1] if c else data[0]


def build_table(data):
    last_date, last_val = data[-1]
    rows = []
    for label, kind, days in LOOKBACKS:
        if kind == "today":
            d, v = last_date, last_val
        elif kind == "milei":
            d, v = nearest_on_or_before(data, MILEI)
        elif kind == "yearend":
            d, v = nearest_on_or_before(data, dt.datetime(last_date.year - 1, 12, 31))
            label = f"Fin {last_date.year - 1}"
        else:  # days
            d, v = nearest_on_or_before(data, last_date - dt.timedelta(days=days))
        delta = last_val - v
        pct = (delta / v) if v else 0.0
        rows.append((label, d, v, delta, pct))
    return last_date, last_val, rows


CHART_PNG = os.path.join(os.path.dirname(os.path.abspath(__file__)), "_grafico.png")


def render_chart(data):
    """Gráfico de la serie desde Milei -> PNG (ejes correctos: fecha en X, u$s en Y)."""
    s = [(d, v) for d, v in data if d >= MILEI]
    xs = [d for d, v in s]
    ys = [v for d, v in s]
    fig, ax = plt.subplots(figsize=(11, 5), dpi=140)
    ax.plot(xs, ys, color="#1F4E78", lw=2.6)
    ax.fill_between(xs, ys, min(ys), color="#1F4E78", alpha=0.08)
    ax.set_title("Depósitos en USD del Sector Privado — BCRA\n(desde la asunción de Milei, 10/12/2023)",
                 fontsize=16, weight="bold")
    ax.set_ylabel("Millones de u$s", fontsize=14)
    ax.tick_params(labelsize=13)
    ax.grid(alpha=0.25)
    ax.xaxis.set_major_locator(mdates.MonthLocator(interval=3))
    ax.xaxis.set_major_formatter(mdates.DateFormatter("%m/%y"))
    last = s[-1]
    ax.annotate(f"{last[1]:,.0f}", xy=last, xytext=(0, 10), textcoords="offset points",
                weight="bold", color="#1F4E78", fontsize=15)
    fig.tight_layout()
    fig.savefig(CHART_PNG)
    plt.close(fig)
    return CHART_PNG


def write_xlsx(data, last_date, last_val, table, mm=None):
    wb = openpyxl.Workbook()

    # ----- Hoja Serie (datos para el gráfico, desde Milei) -----
    ws = wb.active
    ws.title = "Serie"
    ws["A1"] = "Fecha"
    ws["B1"] = "Depósitos USD Sector Privado (millones u$s)"
    for c in ("A1", "B1"):
        ws[c].font = Font(bold=True)
    serie = [(d, v) for d, v in data if d >= MILEI]
    for i, (d, v) in enumerate(serie, start=2):
        ws.cell(row=i, column=1, value=d).number_format = "dd/mm/yyyy"
        ws.cell(row=i, column=2, value=v).number_format = "#,##0"
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 42
    n = len(serie) + 1  # última fila con datos

    # ----- Hoja 1 (gráfico + cuadro) -----
    h = wb.create_sheet("Hoja 1")
    BLUE = "1F4E78"
    title_font = Font(bold=True, size=14, color="FFFFFF")
    title_fill = PatternFill("solid", fgColor=BLUE)
    h["A1"] = "Depósitos en USD del Sector Privado — BCRA"
    h["A1"].font = title_font
    h.merge_cells("A1:E1")
    for col in "ABCDE":
        h[f"{col}1"].fill = title_fill
    h["A2"] = f"Serie diaria, en millones de u$s · Último dato: {last_date.strftime('%d/%m/%Y')} = {last_val:,.0f}"
    h["A2"].font = Font(italic=True, size=10, color="555555")
    h.merge_cells("A2:E2")

    # Gráfico como imagen incrustada (ejes correctos garantizados)
    png = render_chart(data)
    img = XLImage(png)
    img.anchor = "A4"
    h.add_image(img)

    # ----- Cuadro (debajo del gráfico) — layout horizontal de 3 bloques -----
    thin = Side(style="thin", color="BBBBBB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    GREEN, RED = "1F7A1F", "C00000"
    labels = [r[0] for r in table]   # Último, 1 Mes, Fin AAAA, 1 Año, Inicio Milei
    ncols = len(labels)

    def block(r0, title, getval, fmt, colorize):
        h.cell(row=r0, column=1, value=title).font = Font(bold=True, size=11, color=BLUE)
        # fila de encabezados (períodos + fecha de referencia)
        for j, (label, row) in enumerate(zip(labels, table), start=2):
            c = h.cell(row=r0 + 1, column=j, value=label)
            c.font = Font(bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor=BLUE)
            c.alignment = Alignment(horizontal="center"); c.border = border
            fc = h.cell(row=r0 + 2, column=j, value=row[1]); fc.number_format = "dd/mm/yyyy"
            fc.alignment = Alignment(horizontal="center"); fc.font = Font(size=9, color="777777"); fc.border = border
        h.cell(row=r0 + 1, column=1, value="").border = border
        lc = h.cell(row=r0 + 2, column=1, value="Fecha"); lc.font = Font(size=9, italic=True, color="777777"); lc.border = border
        # fila de valores
        vr = r0 + 3
        h.cell(row=vr, column=1, value="Valor").font = Font(bold=True)
        h.cell(row=vr, column=1).border = border
        for j, row in enumerate(table, start=2):
            val = getval(row)
            c = h.cell(row=vr, column=j, value=val); c.number_format = fmt; c.border = border
            c.alignment = Alignment(horizontal="right")
            if colorize and j > 2:  # primer período (Último) no se colorea: Δ=0
                c.font = Font(color=GREEN if val >= 0 else RED)
        return vr

    # Bloque 1: Nivel (debajo de la imagen del gráfico, que ocupa ~filas 4-31)
    r = 34
    block(r, "Depósitos en Dólares del Sector Privado (Mill. de u$s)",
          lambda row: round(row[2]), "#,##0", colorize=False)
    # Bloque 2: Cambio en USD (vs último)
    r += 5
    block(r, "Cambio en los Depósitos (Mill. de u$s)",
          lambda row: round(row[3]), "+#,##0;-#,##0", colorize=True)
    # Bloque 3: Cambio en %
    r += 5
    block(r, "Cambio en los Depósitos (%)",
          lambda row: row[4], "+0.0%;-0.0%", colorize=True)

    for col, w in zip("ABCDEF", [16, 16, 14, 14, 14, 16]):
        h.column_dimensions[col].width = w

    # ----- Bloque Money Market USD (pedido de Fabio S) -----
    r += 5
    if mm:
        mm_fecha, mm_total = mm
        h.cell(row=r, column=1, value="Money Market USD — patrimonio de fondos en u$s (Mill.)").font = Font(bold=True, size=11, color=BLUE)
        rows_mm = [
            ("Depósitos USD (sector privado)", last_val),
            ("Money Market USD (FCI)", mm_total),
            ("Liquidez USD total", last_val + mm_total),
        ]
        for k, (lbl, val) in enumerate(rows_mm):
            rr = r + 1 + k
            cl = h.cell(row=rr, column=1, value=lbl); cl.border = border
            cv = h.cell(row=rr, column=2, value=round(val)); cv.number_format = "#,##0"; cv.border = border
            cv.alignment = Alignment(horizontal="right")
            if lbl.startswith("Liquidez"):
                cl.font = Font(bold=True); cv.font = Font(bold=True)
        h.cell(row=r + 1 + len(rows_mm), column=1,
               value=f"Money Market USD al {mm_fecha} · Fuente: argentinadatos/CAFCI (suma de FCI Mercado de Dinero en dólares).").font = \
            Font(italic=True, size=9, color="888888")
    else:
        h.cell(row=r, column=1,
               value="Money Market USD: no disponible en esta corrida (fuente argentinadatos no respondió).").font = \
            Font(italic=True, size=9, color="888888")

    wb.save(OUT)
    print(f"\nExcel generado: {OUT}")


def main():
    if "--no-download" not in sys.argv or not os.path.exists(SRC):
        download()
    data = load_series()
    print(f"Serie cargada: {len(data)} días, de {data[0][0]:%d/%m/%Y} a {data[-1][0]:%d/%m/%Y}")
    last_date, last_val, table = build_table(data)
    print(f"\nCuadro (último dato {last_date:%d/%m/%Y} = {last_val:,.0f} M u$s):")
    for label, d, v, delta, pct in table:
        print(f"  {label:30} {d:%d/%m/%Y}  {v:>9,.0f}   Δ {delta:>+9,.0f}  ({pct:>+6.1%})")
    mm = fetch_mm_usd()
    if mm:
        print(f"\nMoney Market USD (al {mm[0]}): {mm[1]:,.0f} M u$s  ->  Liquidez USD total: {last_val + mm[1]:,.0f} M")
    write_xlsx(data, last_date, last_val, table, mm)


if __name__ == "__main__":
    main()
